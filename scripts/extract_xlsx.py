"""Extract worksheet data, images, and chart metadata from XLSX into markdown.

Usage:
	python extract_xlsx.py <input.xlsx> <output_dir>
"""

from __future__ import annotations

import io
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


def usage() -> None:
	print("Usage: python extract_xlsx.py <input.xlsx> <output_dir>")


def sanitize(value: str) -> str:
	cleaned = re.sub(r"[^A-Za-z0-9._ -]", "_", value.strip())
	return cleaned or "workbook"


def escape_md(value: str) -> str:
	return value.replace("|", "\\|").replace("\n", " ").strip()


def normalize_cell(value: object) -> str:
	if value is None:
		return ""
	if isinstance(value, (int, float)):
		return str(value)
	return str(value)


def rows_to_markdown(rows: list[list[str]]) -> str:
	if not rows:
		return ""

	width = max(len(row) for row in rows)
	padded = [row + [""] * (width - len(row)) for row in rows]
	header = [escape_md(cell) or " " for cell in padded[0]]
	divider = ["---"] * width
	lines = [
		"| " + " | ".join(header) + " |",
		"| " + " | ".join(divider) + " |",
	]

	for row in padded[1:]:
		lines.append("| " + " | ".join(escape_md(cell) or " " for cell in row) + " |")

	return "\n".join(lines)


def trim_matrix(matrix: list[list[str]]) -> list[list[str]]:
	if not matrix:
		return matrix

	# Remove fully empty trailing rows.
	while matrix and not any(cell.strip() for cell in matrix[-1]):
		matrix.pop()

	if not matrix:
		return matrix

	# Determine the last non-empty column across all rows.
	last_col = 0
	for row in matrix:
		for idx, cell in enumerate(row, start=1):
			if cell.strip():
				last_col = max(last_col, idx)

	if last_col == 0:
		return []

	return [row[:last_col] for row in matrix]


def extract_sheet_rows(sheet, max_rows: int = 1000) -> tuple[list[list[str]], bool]:
	data: list[list[str]] = []
	truncated = False

	upper = min(sheet.max_row, max_rows)
	for row in sheet.iter_rows(min_row=1, max_row=upper, min_col=1, max_col=sheet.max_column):
		data.append([normalize_cell(cell.value) for cell in row])

	if sheet.max_row > max_rows:
		truncated = True

	return trim_matrix(data), truncated


def save_sheet_images(sheet, images_dir: Path, sheet_slug: str) -> list[str]:
	image_paths: list[str] = []
	for index, image in enumerate(getattr(sheet, "_images", []), start=1):
		raw = None
		if hasattr(image, "_data"):
			raw = image._data()
		if raw is None:
			continue

		out_name = f"{sheet_slug}_image_{index:02d}.png"
		out_path = images_dir / out_name
		try:
			with Image.open(io.BytesIO(raw)) as img:
				img.save(out_path)
		except Exception:
			out_path.write_bytes(raw)

		image_paths.append(out_name)

	return image_paths


def chart_title(chart: object, fallback_index: int) -> str:
	title_obj = getattr(chart, "title", None)
	if title_obj is None:
		return f"Chart {fallback_index}"
	text_obj = getattr(title_obj, "tx", None)
	if text_obj and getattr(text_obj, "rich", None):
		parts = []
		for paragraph in text_obj.rich.p:
			for run in paragraph.r:
				if run.t:
					parts.append(str(run.t))
		if parts:
			return "".join(parts)
	return f"Chart {fallback_index}"


def main() -> int:
	if len(sys.argv) != 3:
		usage()
		return 1

	input_path = Path(sys.argv[1]).expanduser().resolve()
	output_dir = Path(sys.argv[2]).expanduser().resolve()

	if not input_path.exists() or input_path.suffix.lower() != ".xlsx":
		print(f"Invalid XLSX input file: {input_path}")
		return 1

	output_dir.mkdir(parents=True, exist_ok=True)
	images_dir = output_dir / "images"
	images_dir.mkdir(parents=True, exist_ok=True)
	markdown_path = output_dir / f"{sanitize(input_path.stem)}.md"

	workbook = load_workbook(input_path, data_only=True)
	lines: list[str] = [f"# {input_path.name}", "", f"Sheets: {len(workbook.worksheets)}", ""]

	for sheet in workbook.worksheets:
		sheet_slug = sanitize(sheet.title).replace(" ", "_")
		lines.append(f"## Sheet: {sheet.title}")
		lines.append("")

		rows, truncated = extract_sheet_rows(sheet)
		if rows:
			lines.append(rows_to_markdown(rows))
			lines.append("")
		else:
			lines.append("_No tabular cell data extracted._")
			lines.append("")

		if truncated:
			lines.append("_Note: Row extraction truncated at 1000 rows for performance._")
			lines.append("")

		images = save_sheet_images(sheet, images_dir, sheet_slug)
		if images:
			lines.append("### Embedded Images")
			lines.append("")
			for image_name in images:
				lines.append(f"- ![Embedded image](./images/{image_name})")
			lines.append("")

		charts = getattr(sheet, "_charts", [])
		if charts:
			lines.append("### Charts")
			lines.append("")
			for chart_index, chart in enumerate(charts, start=1):
				lines.append(f"- {chart_title(chart, chart_index)}")
			lines.append("")

	markdown_path.write_text("\n".join(lines), encoding="utf-8")
	print(f"Extracted: {input_path}")
	print(f"Markdown: {markdown_path}")
	print(f"Images: {images_dir}")
	return 0


if __name__ == "__main__":
	raise SystemExit(main())
