"""
Extract data, embedded images, and chart information from Excel (.xlsx) files.
Each sheet becomes a section with data as markdown tables, embedded images saved,
and charts rendered via Excel COM automation on Windows.

Usage: python extract_xlsx.py <input.xlsx> <output_dir>
"""

import sys
import os
from pathlib import Path


def render_charts_via_com(input_path: str, images_dir: Path) -> dict[str, list[str]]:
    """Use Excel COM to export chart objects as images. Returns {sheet_name: [image_paths]}."""
    abs_path = str(Path(input_path).resolve())
    chart_images: dict[str, list[str]] = {}

    try:
        import comtypes.client

        excel = comtypes.client.CreateObject("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_path, ReadOnly=True)

        chart_count = 0
        for ws in wb.Worksheets:
            sheet_charts = []
            for chart_obj in ws.ChartObjects():
                chart_count += 1
                chart_name = chart_obj.Name or f"chart{chart_count}"
                safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in chart_name)
                image_name = f"chart_{ws.Name}_{safe_name}.png"
                image_path = str(images_dir / image_name)
                try:
                    chart_obj.Chart.Export(image_path, "PNG")
                    sheet_charts.append(image_name)
                except Exception:
                    pass
            if sheet_charts:
                chart_images[ws.Name] = sheet_charts

        wb.Close(SaveChanges=False)
        excel.Quit()
        if chart_count > 0:
            print(f"Rendered {chart_count} charts via Excel COM")
        return chart_images

    except Exception as e:
        print(f"Excel COM chart rendering unavailable ({e}), trying win32com...")

    try:
        import win32com.client

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_path, ReadOnly=True)

        chart_count = 0
        for ws in wb.Worksheets:
            sheet_charts = []
            for chart_obj in ws.ChartObjects():
                chart_count += 1
                chart_name = chart_obj.Name or f"chart{chart_count}"
                safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in chart_name)
                image_name = f"chart_{ws.Name}_{safe_name}.png"
                image_path = str(images_dir / image_name)
                try:
                    chart_obj.Chart.Export(image_path, "PNG")
                    sheet_charts.append(image_name)
                except Exception:
                    pass
            if sheet_charts:
                chart_images[ws.Name] = sheet_charts

        wb.Close(SaveChanges=False)
        excel.Quit()
        if chart_count > 0:
            print(f"Rendered {chart_count} charts via win32com")
        return chart_images

    except Exception as e:
        print(f"win32com chart rendering also unavailable ({e})")
        return chart_images


def extract_xlsx(input_path: str, output_dir: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True)
    output_path = Path(output_dir)
    images_dir = output_path / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    filename = Path(input_path).stem

    # --- Phase 1: Render charts via COM ---
    chart_images = render_charts_via_com(input_path, images_dir)

    # --- Phase 2: Extract data, images, and chart info ---
    md_lines = [f"# Extracted: {filename}\n"]
    md_lines.append(f"**Source**: {Path(input_path).name}\n")
    md_lines.append(f"**Sheets**: {len(wb.sheetnames)}\n")
    md_lines.append("---\n")

    image_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_lines.append(f"\n## Sheet: {sheet_name}\n")

        # --- Embedded images ---
        if hasattr(ws, '_images') and ws._images:
            for img in ws._images:
                try:
                    image_count += 1
                    ext = "png"
                    if hasattr(img, 'format') and img.format:
                        ext = img.format.lower()
                    image_name = f"sheet_{sheet_name}_img{image_count}.{ext}"
                    image_path = images_dir / image_name
                    with open(image_path, "wb") as f:
                        f.write(img._data())
                    anchor_info = ""
                    if hasattr(img, 'anchor') and img.anchor:
                        anchor_info = f" (at {img.anchor})"
                    md_lines.append(f"![Embedded image{anchor_info}](./images/{image_name})\n")
                except Exception:
                    pass

        # --- Chart renders from COM ---
        if sheet_name in chart_images:
            for chart_img in chart_images[sheet_name]:
                md_lines.append(f"![Chart](./images/{chart_img})\n")

        # --- Chart objects detected via openpyxl (note if COM didn't render) ---
        if hasattr(ws, '_charts') and ws._charts and sheet_name not in chart_images:
            md_lines.append(f"\n*{len(ws._charts)} chart(s) detected but could not be rendered (Excel not available). Open the file in Excel to view.*\n")

        # --- Table data ---
        if ws.max_row is None or ws.max_row == 0:
            md_lines.append("*Empty sheet*\n")
            continue

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            md_lines.append("*Empty sheet*\n")
            continue

        # Filter out completely empty rows
        non_empty = [row for row in rows if any(cell is not None for cell in row)]
        if not non_empty:
            md_lines.append("*Empty sheet*\n")
            continue

        col_count = len(rows[0])

        # First row as header
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        md_lines.append("| " + " | ".join(headers) + " |")
        md_lines.append("| " + " | ".join(["---"] * col_count) + " |")

        # Data rows (limit to first 100 rows for readability)
        max_rows = min(len(rows), 101)
        for row in rows[1:max_rows]:
            cells = []
            for cell in row[:col_count]:
                value = str(cell) if cell is not None else ""
                value = value.replace("|", "\\|").replace("\n", " ")
                cells.append(value)
            md_lines.append("| " + " | ".join(cells) + " |")

        if len(rows) > 101:
            md_lines.append(f"\n*... {len(rows) - 101} additional rows truncated*\n")

        md_lines.append(f"\n**Rows**: {len(rows) - 1} (excluding header) | **Columns**: {col_count}\n")

    # Write markdown
    md_path = output_path / f"{filename}.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))

    chart_total = sum(len(v) for v in chart_images.values())
    print(f"Extracted {len(wb.sheetnames)} sheets, {image_count} embedded images, {chart_total} chart renders")
    print(f"Output: {md_path}")
    return str(md_path)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python extract_xlsx.py <input.xlsx> <output_dir>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_directory = sys.argv[2]

    if not os.path.exists(input_file):
        print(f"Error: File not found: {input_file}")
        sys.exit(1)

    extract_xlsx(input_file, output_directory)
